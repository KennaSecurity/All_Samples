import requests
import pandas as pd
import json
import sys
import io
import os
import datetime
import gzip
import jsonlines
from tqdm import tqdm
from pandas import json_normalize
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def flatten_json(nested_json, exclude=['roles']):
    out = {}
    def flatten(x, name='', exclude=exclude):
        if type(x) is dict:
            for a in x:
                if a not in exclude: 
                    flatten(x[a], name + a + '_')
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + '_')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(nested_json)
    return out

token = os.environ.get('API_KEY')
base_url = "https://api.kennasecurity.com"
users_url= base_url + "/users"
roles_url = base_url + "/roles"
audit_logs_url = base_url + "/audit_logs/"

headers = {"Accept": "application/json", "X-Risk-Token":token}

users_response = requests.get(users_url, headers=headers).json()

users_df = pd.DataFrame(json_normalize([flatten_json(x) for x in users_response['users']]))
users_df = users_df.rename(columns={"id":"user_id","created_at":"user_created_at","updated_at":"user_updated_at"})

users_df['user_created_at'] = pd.to_datetime(users_df['user_created_at'], format='%Y-%m-%d', errors='coerce').dt.date
users_df['last_sign_in_at'] = pd.to_datetime(users_df['last_sign_in_at'], format='%Y-%m-%d', errors='coerce').dt.date

roles_response = requests.get(roles_url, headers=headers).json()

roles_df = pd.DataFrame(json_normalize([flatten_json(x) for x in roles_response['roles']]))
roles_df = roles_df.rename(columns={"id": "role_id","created_at":"role_created_at","updated_at":"role_updated_at","name":"role_name"})

# Save 'Users' and 'Roles' data to Excel and apply conditional formatting
with pd.ExcelWriter('cvm_user_audit.xlsx', engine='xlsxwriter', date_format='m/d/yyyy') as writer:
    users_df.to_excel(writer, sheet_name='Users')
    roles_df.to_excel(writer, sheet_name='Roles')

    workbook = writer.book
    red_format = workbook.add_format({'bg_color': '#FFC7CE','font_color': '#9C0006'})
    yellow_format = workbook.add_format({'bg_color': '#FFEB9C','font_color': '#9C6500'})

    users_sheet = writer.sheets['Users']

    users_sheet.conditional_format('$J$2:$J$99999', {'type': 'blanks', 'format': red_format})
    users_sheet.conditional_format('$J$2:$J$99999', {'type': 'formula', 'criteria': '=J2<TODAY()-30', 'format': yellow_format})

# Fetch audit logs data
params = {
    "start_date": "2024-01-03T00:00:00",
    "end_date": "2024-01-06T00:00:00",
}

# Update headers for audit_logs_url
headers = {
    "Accept": "application/gzip",
    "X-Risk-Token": token,
    "User-Agent": 'user_audit/1.0.0 (Kenna Security)'
}

print('Fetching audit logs data...')
response = requests.get(audit_logs_url, headers=headers, params=params, timeout=60, stream=True)

print('Content-Encoding:', response.headers.get('Content-Encoding'))
print('Receiving data...')

gzip_content = b""
for chunk in tqdm(response.iter_content(chunk_size=8192)):  # Wrap the iterable with tqdm()
    if chunk:  # filter out keep-alive new chunks
        gzip_content += chunk

gzip_file = io.BytesIO(gzip_content)

print('Request sent. Decompressing response...')
with gzip.GzipFile(fileobj=gzip_file) as f:
    audit_logs_data = f.read().decode()

# Create a StringIO object from the string
audit_logs_io = io.StringIO(audit_logs_data)

# Use the jsonlines library to read the data
print('Response decompressed. Parsing JSON...')
audit_logs = []
with jsonlines.Reader(audit_logs_io) as reader:
    for i, obj in enumerate(reader.iter(type=dict, skip_invalid=True), start=1):
        #print(f"Line {i}: {obj}")  # print line number and JSON object
        audit_logs.append(obj)

print('JSON parsed. Processing data...')
audit_logs_processed = []
for log in audit_logs:
    details = log.get("audit_log_event", {}).get("details", {})
    url = details.get("url")
    source = details.get("source")
    # Ignore this entry if 'url' is '{base_url}/reports/sla_adherences' and 'source' is 'API'
    if url == f"{base_url}/reports/sla_adherences" and source == 'API':
        continue
    event_data = {
        "kenna_user_id": log.get("audit_log_event", {}).get("kenna_user_id"),
        "user_email": log.get("audit_log_event", {}).get("user_email"),
        "source": source,
        "http_method": details.get("http_method"),
        "url": url,
        "name": log.get("audit_log_event", {}).get("name"),
    }
    audit_logs_processed.append(event_data)

audit_df = pd.DataFrame(audit_logs_processed)

# Open the existing workbook with 'openpyxl'
wb = load_workbook('cvm_user_audit.xlsx')

# Write the 'Audit Logs' DataFrame to the workbook
with pd.ExcelWriter('cvm_user_audit.xlsx', engine='openpyxl') as writer:
    writer.book = wb
    merged_df = pd.merge(users_df, audit_df, left_on="user_id", right_on="kenna_user_id", how="inner")
    merged_df.to_excel(writer, sheet_name='Audit Logs')

    # Get the 'Users' sheet
    users_sheet = wb['Users']

    # Define a fill for highlighting cells
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    # Get emails from audit logs data with 'source' as 'API'
    api_emails = [log['user_email'] for log in audit_logs_processed if log['source'] == 'API']

    # Iterate over the rows in the 'Users' DataFrame
    for i, row in users_df.iterrows():
        email = row['email']  # Assuming 'email' is a column in your DataFrame
        # Check if the email is in api_emails
        if email in api_emails:
            # If it is, highlight the entire row
            for j in range(1, len(row) + 1):
                users_sheet.cell(row=i+2, column=j).fill = green_fill  # i+2 because DataFrame is 0-indexed and Worksheet is 1-indexed, and we have a header row

    # Save the workbook
    wb.save('cvm_user_audit.xlsx')

print('User, role, and audit log data has been saved to the file cvm_user_audit.xlsx.')
print('Users that have never logged in are highlighted in red. Users that have not logged in for over 30 days are highlighted in yellow.')
print('Users that have "source" as "API" in "Audit Logs" and used their key in the stipulated audit period are highlighted in green.')